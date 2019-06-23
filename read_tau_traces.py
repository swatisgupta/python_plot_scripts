#!/usr/bin/python

from __future__ import division 
from ctypes import *
import sys
import math
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy
from statistics import mean
import matplotlib.dates as mdates
import datetime as dt
import pandas as pd
from pandas import Series
import os.path
from pandas import Series
from functools import partial
from pandas.tseries.frequencies import to_offset
from pandas import Timestamp
from openpyxl import load_workbook
import xlsxwriter
import os
import shutil

ini_timestamp=0
sample_freq='1S'
frequency=1
max_duration=90

is_microsecond=1 

def copy_rename(old_file_name, new_file_name, o_dir, n_dir):
    shutil.copy(old_file_name, n_dir)
    dst_file = os.path.join(n_dir, old_file_name)
    new_dst_file_name = os.path.join(n_dir, new_file_name)
    os.rename(dst_file, new_dst_file_name)

def create_workbook(filename, sheetname):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet(sheetname)
    workbook.close()

class TAU_EV32(Structure):
    _fields_ = [ ('ev', c_int),
                 ('nid', c_ushort ), #unsigned short),
                 ('tid', c_ushort ), #unsigned short),
                 ( 'par', c_ulonglong),
                 ( 'ti', c_ulonglong ) ]

Colors = ['blue', 'orange', 'green', 'red', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan', 'blue', 'orange']

function_id = {}
tag = {}
unmap = {}
umpi = {}
mylist = {}

L3 = 0 #Last level cache misses
LD = 1 #load stalls
ST = 2 #store stalls
TCYC = 3 #unhalted cycles
INS = 4 #instructions retired
MEMO = 5 #total memory operations
VHWM = 6 #virtual memory size or swap space
RSS = 7 #resident set size
TCA = 8 #total last level cache references


MEM1_RD1 = 9 #reads completed or issued to queue
MEM1_WR1 = 10 #writes completed oor issued to queue

MEM1_RD2 = 11 #reads completed or issued to queue
MEM1_WR2 = 12 #writes completed oor issued to queue

MEM1_RD3 = 13 #reads completed or issued to queue
MEM1_WR3 = 14 #writes completed oor issued to queue

MEM1_RD4 = 15 #reads completed or issued to queue
MEM1_WR4 = 16 #writes completed oor issued to queue

MEM2_RD1 = 17 #reads completed or issued to queue
MEM2_WR1 = 18 #writes completed oor issued to queue

MEM2_RD2 = 19 #reads completed or issued to queue
MEM2_WR2 = 20 #writes completed oor issued to queue

MEM2_RD3 = 21 #reads completed or issued to queue
MEM2_WR3 = 22 #writes completed oor issued to queue

MEM2_RD4 = 23 #reads completed or issued to queue
MEM2_WR4 = 24 #writes completed oor issued to queue

RQ_RD1 = 25 #total reads requested
RQ_WR1 = 26 #total writes requested
RQ_RD2 = 27 #total reads requested
RQ_WR2 = 28 #total writes requested

IR_RD1 = 29 #total reads requested
IR_WR1 = 30 #total writes requested
IR_RD2 = 31 #total reads requested
IR_WR2 = 32 #total writes requested
DR_A = 33
DR_O =  34
SQ_F = 35
CMPI = 35
MPIS = 36
MPIR = 37
MPISR = 38

counter_name = {L3:'L3', TCA:'TCA', INS:'INS', LD:'LD', ST:'ST', TCYC:'TCYC', MEMO:'MEMO', RSS:'RSS', VHWM:'VHWM', MEM1_RD1:'MEM1_RD1', MEM1_RD2:'MEM1_RD2', MEM1_RD3:'MEM1_RD3', MEM1_RD4:'MEM1_RD4', MEM1_WR1:'MEM1_WR1', MEM1_WR2:'MEM1_WR2', MEM1_WR3:'MEM1_WR3', MEM1_WR4:'MEM1_WR4', MEM2_RD1:'MEM2_RD1', MEM2_RD2:'MEM2_RD2', MEM2_RD3:'MEM2_RD3', MEM2_RD4:'MEM1_RD4', MEM2_WR1:'MEM2_WR1', MEM2_WR2:'MEM2_WR2', MEM2_WR3:'MEM2_WR3', MEM2_WR4:'MEM2_WR4', RQ_WR1:'RQ_WR1', RQ_WR2:'RQ_WR2', RQ_RD1:'RQ_RD1', RQ_RD2:'RD_RD2', IR_WR1:'IR_WR1', IR_WR2:'IR_WR2', IR_RD1:'IR_RD1', IR_RD2:'IR_RD2',MPIS:'MPIS', MPIR:'MPIR', MPISR:'MPISR', DR_A:'DR_A', DR_O:'DR_O', SQ_F:'SQ_F'}    

unmap[L3] = -1
unmap[TCA] = -1
unmap[INS] = -1
unmap[LD] = -1
unmap[ST] = -1
unmap[TCYC] = -1
unmap[MEMO] = -1

unmap[MEM1_RD1] = -1
unmap[MEM1_RD2] = -1
unmap[MEM1_RD3] = -1
unmap[MEM1_RD4] = -1

unmap[MEM1_WR1] = -1
unmap[MEM1_WR2] = -1
unmap[MEM1_WR3] = -1
unmap[MEM1_WR4] = -1

unmap[MEM2_RD1] = -1
unmap[MEM2_RD2] = -1
unmap[MEM2_RD3] = -1
unmap[MEM2_RD4] = -1

unmap[MEM2_WR1] = -1
unmap[MEM2_WR2] = -1
unmap[MEM2_WR3] = -1
unmap[MEM2_WR4] = -1

unmap[RQ_RD1] = -1
unmap[RQ_RD2] = -1

unmap[RQ_WR1] = -1
unmap[RQ_WR2] = -1
unmap[IR_RD1] = -1
unmap[IR_RD2] = -1

unmap[IR_WR1] = -1
unmap[IR_WR2] = -1

unmap[DR_A] = -1
unmap[DR_O] = -1
unmap[SQ_F] = -1

unmap[VHWM] = -1
unmap[RSS] = -1

umpi[MPIS%CMPI] = []
umpi[MPIR%CMPI] = []
umpi[MPISR%CMPI] = []


def get_counters_info(system):
    counters = {}
    if system == "deepthought2":
        counters[L3]=["PAPI_NATIVE_LAST_LEVEL_CACHE_MISSES", "LIKWID_L3_LAT_CACHE_MISS_PMC1"]
        counters[INS]=["PAPI_NATIVE_INSTRUCTIONS_RETIRED", "LIKWID_INSTR_RETIRED_ANY_FIXC0"]
        counters[TCA]=["PAPI_NATIVE_LAST_LEVEL_CACHE_REFERENCES", "LIKWID_L3_LAT_CACHE_REFERENCE_PMC0"]
        counters[LD]=["PAPI_NATIVE_CYCLE_ACTIVITY_STALLS_LDM_PENDING", "PAPI_NATIVE_UOPS_RETIRED_STALL_CYCLES", "LIKWID_CYCLE_ACTIVITY_CYCLES_LDM_PENDING_PMC3"]
        counters[ST]=["PAPI_NATIVE_RESOURCE_STALLS_SB", "LIKWID_RESOURCE_STALLS_SB_PMC4"]
        counters[TCYC]=["PAPI_NATIVE_perf__CPU-CYCLES", "PAPI_NATIVE_ix86arch__UNHALTED_CORE_CYCLES"i, "LIKWID_CPU_CLK_UNHALTED_CORE_FIXC1"]
        counters[MEMO]=["PAPI_NATIVE_MEM_LOAD_UOPS_LLC_MISS_RETIRED", "LIKWID_MEM_UOPS_RETIRED_LOADS_PMC2"]
        counters[MEM1_WR1]=["LIKWID_WPQ_INSERTS_MBOX0C1_EDGEDETECT_THRESHOLD=0x1", "LIKWID_CAS_COUNT_WR_MBOX0C1_EDGEDETECT_THRESHOLD=0x1", "LIKWID_WPQ_CYCLES_NE_MBOX0C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM1_WR2]=["LIKWID_WPQ_INSERTS_MBOX1C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_WR_MBOX1C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX1C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM1_WR3]=["LIKWID_WPQ_INSERTS_MBOX2C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_WR_MBOX2C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX2C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM1_WR4]=["LIKWID_WPQ_INSERTS_MBOX3C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_WR_MBOX3C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX3C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM1_RD1]=["LIKWID_RPQ_INSERTS_MBOX0C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX0C0" , "LIKWID_RPQ_CYCLES_NE_MBOX0C0_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM1_RD2]=["LIKWID_RPQ_INSERTS_MBOX1C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX1C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_RPQ_CYCLES_NE_MBOX1C0_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM1_RD3]=["LIKWID_RPQ_INSERTS_MBOX2C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX2C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_RPQ_CYCLES_NE_MBOX2C0_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM1_RD4]=["LIKWID_RPQ_INSERTS_MBOX3C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX3C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX3C0_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM2_WR1]=["LIKWID_WPQ_INSERTS_MBOX4C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_WR_MBOX4C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX4C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM2_WR2]=["LIKWID_WPQ_INSERTS_MBOX5C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_WR_MBOX5C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX5C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM2_WR3]=["LIKWID_WPQ_INSERTS_MBOX6C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_WR_MBOX6C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX6C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM2_WR4]=["LIKWID_WPQ_INSERTS_MBOX7C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_WR_MBOX7C1_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX7C1_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM2_RD1]=["LIKWID_RPQ_INSERTS_MBOX4C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX4C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_RPQ_CYCLES_NE_MBOX4C0_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM2_RD2]=["LIKWID_RPQ_INSERTS_MBOX5C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX5C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_RPQ_CYCLES_NE_MBOX5C0_EDGEDETECT_THRESHOLD=0x1"]
        counters[MEM2_RD3]=["LIKWID_RPQ_INSERTS_MBOX6C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX6C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_RPQ_CYCLES_NE_MBOX6C0_EDGEDETECT_THRESHOLD=0x1"]        
        counters[MEM2_RD4]=["LIKWID_RPQ_INSERTS_MBOX7C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_CAS_COUNT_RD_MBOX7C0_EDGEDETECT_THRESHOLD=0x1" , "LIKWID_WPQ_CYCLES_NE_MBOX7C0_EDGEDETECT_THRESHOLD=0x1"]        
        counters[RQ_RD1]=["LIKWID_REQUESTS_READS_BBOX0C0" , "LIKWID_IMC_READS_NORMAL_BBOX0C0_EDGEDETECT_THRESHOLD=0x1"]        
        counters[RQ_RD2]=["LIKWID_REQUESTS_READS_BBOX2C0" , "LIKWID_IMC_READS_NORMAL_BBOX1C0_EDGEDETECT_THRESHOLD=0x1"]        
        counters[RQ_WR1]=["LIKWID_REQUESTS_WRITES_BBOX0C1" , "LIKWID_IMC_WRITES_ALL_BBOX0C1_EDGEDETECT_THRESHOLD=0x1"]        
        counters[RQ_WR2]=["LIKWID_REQUESTS_WRITES_BBOX1C1" , "LIKWID_IMC_WRITES_ALL_BBOX1C1_EDGEDETECT_THRESHOLD=0x1"]        
        counters[IR_WR1]=["LIKWID_IMC_WRITES_ALL_BBOX0C3"]        
        counters[IR_WR2]=["LIKWID_IMC_WRITES_ALL_BBOX1C3"]        
        counters[IR_RD1]=["LIKWID_IMC_READS_NORMAL_BBOX0C2"]        
        counters[IR_RD2]=["LIKWID_IMC_READS_NORMAL_BBOX1C2"]        
        counters[IR_DR_O]=["PAPI_NATIVE_OFFCORE_REQUESTS_OUTSTANDING_DEMAND_DATA_RD"]        
        counters[IR_DR_A]=["PAPI_NATIVE_OFFCORE_REQUESTS_DEMAND_DATA_RD"]        
        counters[SQ_F]=["PAPI_NATIVE_OFFCORE_REQUESTS_BUFFER_SQ_FULL"]        
     elif system == "summit":
        counters[L3]=["PAPI_NATIVE_LAST_LEVEL_CACHE_MISSES"]
        counters[INS]=["PAPI_NATIVE_INSTRUCTIONS_RETIRED"]
        counters[TCA]=["PAPI_NATIVE_LAST_LEVEL_CACHE_REFERENCES"]
        counters[LD]=["PAPI_NATIVE_CYCLE_ACTIVITY_STALLS_LDM_PENDING", "PAPI_NATIVE_UOPS_RETIRED_STALL_CYCLES"]
        counters[ST]=["PAPI_NATIVE_RESOURCE_STALLS_SB"]
        counters[TCYC]=["PAPI_NATIVE_perf__CPU-CYCLES", "PAPI_NATIVE_ix86arch__UNHALTED_CORE_CYCLES"]
        counters[MEMO]=["PAPI_NATIVE_MEM_LOAD_UOPS_LLC_MISS_RETIRED"]
     else:
        counters[L3]=[""]
        counters[INS]=[""]
        counters[TCA]=[""]
        counters[LD]=[""]
        counters[ST]=[""]
        counters[TCYC]=[""]
     return counters

def Average(lst): 
    sumlst = 0
    for val in lst:
        sumlst += val
    return sumlst / len(lst)

def get_event(event_id):
    if event_id == unmap[L3]:
        return L3
    elif event_id == unmap[INS]:
        return INS
    elif event_id == unmap[LD]:
        return LD
    elif event_id == unmap[ST]:
        return ST
    elif event_id == unmap[TCYC]:
        return TCYC
    elif event_id == unmap[TCA]:
        return TCA
    elif event_id == unmap[VHWM]:
        return VHWM
    elif event_id == unmap[RSS]:
        return RSS
    elif event_id == unmap[MEMO]:
        return MEMO
    elif event_id == unmap[MEM1_RD1]:
        return MEM1_RD1
    elif event_id == unmap[MEM1_RD2]:
        return MEM1_RD2
    elif event_id == unmap[MEM1_RD3]:
        return MEM1_RD3
    elif event_id == unmap[MEM1_RD4]:
        return MEM1_RD4
    elif event_id == unmap[MEM1_WR1]:
        return MEM1_WR1
    elif event_id == unmap[MEM1_WR2]:
        return MEM1_WR2
    elif event_id == unmap[MEM1_WR3]:
        return MEM1_WR3
    elif event_id == unmap[MEM1_WR4]:
        return MEM1_WR4
    elif event_id == unmap[MEM2_RD1]:
        return MEM2_RD1
    elif event_id == unmap[MEM2_RD2]:
        return MEM2_RD2
    elif event_id == unmap[MEM2_RD3]:
        return MEM2_RD3
    elif event_id == unmap[MEM2_RD4]:
        return MEM2_RD4
    elif event_id == unmap[MEM2_WR1]:
        return MEM2_WR1
    elif event_id == unmap[MEM2_WR2]:
        return MEM2_WR2
    elif event_id == unmap[MEM2_WR3]:
        return MEM2_WR3
    elif event_id == unmap[MEM2_WR4]:
        return MEM2_WR4
    elif event_id == unmap[RQ_WR1]:
        return RQ_WR1
    elif event_id == unmap[RQ_WR2]:
        return RQ_WR2
    elif event_id == unmap[RQ_RD1]:
        return RQ_RD1
    elif event_id == unmap[RQ_RD2]:
        return RQ_RD2
    elif event_id == unmap[IR_WR1]:
        return IR_WR1
    elif event_id == unmap[IR_WR2]:
        return IR_WR2
    elif event_id == unmap[IR_RD1]:
        return IR_RD1
    elif event_id == unmap[IR_RD2]:
        return IR_RD2
    elif event_id == unmap[DR_A]:
        return DR_A
    elif event_id == unmap[DR_O]:
        return DR_O
    elif event_id == unmap[SQ_F]:
        return SQ_F
    elif umpi[MPIS%CMPI].count(event_id) > 0:
        return MPIS
    elif umpi[MPIR%CMPI].count(event_id) > 0:
        return MPIR
    elif umpi[MPISR%CMPI].count(event_id) > 0:
        return MPISR
    else:
        return -1
 
def get_trigger(word, counters):
    all_counters = counters.keys()
    trigger = -1
    for key in all_counters:    
        if word in counters[key]:
            trigger = key
            print("Found " + word)
            return trigger

    if word == "MPI_Gather()  ":
        trigger = MPISR
        print("Found MPI_Gather()  ")
    elif word == "MPI_Gatherv()  ":
        trigger = MPISR
        print("Found MPI_Gatherv()  ")
    elif word == "MPI_Bcast()  ":
        trigger = MPISR
        print("Found MPI_Bcast()  ")
    elif word == "MPI_Allreduce()  ":
        trigger = MPISR
        print("Found MPI_Allreduce()  ")
    elif word == "MPI_Allgather()  ":
        trigger = MPISR
        print("Found MPI_Allgather()  ")
    elif word == "MPI_Allgatherv()  ":
        trigger = MPISR
        print("Found MPI_Allgatherv()  ")
    elif word == "MPI_Send()  ":
        trigger = MPIS
        print("Found MPI_Send()  ")
    elif word == "MPI_Recv()  ":
        trigger = MPIR
        print("Found MPI_Recv()  ")
    elif word == "MPI_Isend()  ":
        trigger = MPIS
        print("Found MPI_Isend()  ")
    elif word == "MPI_Irecv()  ":
        trigger = MPIR
        print("Found MPI_Irecv()  ")
    elif word == "Peak Memory Usage Resident Set Size (VmHWM) (KB)" :
        trigger = VHWM
        print("Found VHWM")
    elif word == "Memory Footprint (VmRSS) (KB)":
        trigger = RSS
        print("Found Memory Footprint (VmRSS) (KB)")
    else:
        trigger = -1
    return trigger
	  
def read_edf(name, system):
    counters=[]
    avail_counters = get_counters_info(system)
    with open(name, 'r') as f:
        line = f.readline()        
        line = f.readline()        
        for line in f:
            words = line.split('\"')
            trigger = get_trigger(words[1], avail_counters)
            if ( trigger != -1 ) :
                words2 = words[0].split(' ')
                counters.append(trigger)
                function_id[int(words2[0])] = trigger
                if trigger == MPIR or trigger == MPIS or trigger == MPISR:
                    umpi[trigger%CMPI].append(int(words2[0]))
                else: 
                    unmap[trigger] = int(words2[0])
    return counters


def read_trace(tracedir, nprocs, nthreads, reference):
    i = 0
    while i < nprocs:
        name = tracedir + "/tautrace." + str(i + reference) + ".0.0.trc"     
        if os.path.exists(name) == False: 
            print("TRace: reading failed for ", name) 
            name = tracedir + "/tautrace.0.0.0.trc"     
        dicts = [{}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}]
        with open(name, 'rb') as f:
            event = TAU_EV32()
            while True:
            #for chunk in iter(partial(f.read, 192), ''):
                nrd = f.readinto(event)
                if nrd == 0:
                    break
                #print ("Event is" , event.ev, " nprocs is ", event.nid, " thread is ", event.tid, " parameter id is", event.par, " timestamp is", event.ti  )        
                #print(event.nid)        
                #print(event.tid)        
                ev_idx = get_event(event.ev)
                #print("Event is ", event.ev, " Index is ", ev_idx) 
                if event.nid == (i + reference) and ev_idx != -1:
                    dicts[ev_idx][event.ti] = event.par
        mylist[i] = dicts           
        i += 1

def getAllKeys(dict, nths, index):
    i = 0

    all_keys = {}
    while i < nths:
        #print("Pr", i)
        key_vals = []
        dictth = dict[i]
        flag = 0
        keys = sorted(dict[i][index])
        for key in keys:
            if dictth[index][key] == 0 and (index == VHWM or index == RSS):
                flag= flag + 1
                if flag == 2:
                   key_vals.append(key)
                   flag = -1
                continue
            else:
                flag = -1    
                key_vals.append(key)
        all_keys[i] = key_vals
        i+=1
    return all_keys

def get_values_for(dict, nths, index, keys, isdiff=0):
    i = 0
    max_val = 0
    all_vals = {}
    new_keys = {}
    while i < nths:
        #print("Pr" , i)
        vals = []
        val_l1=[]
        dictth = dict[i]
        for key in keys[i]:
            if dictth[index].get(key) == None or dictth[index][key] == 0 and (index == VHWM or index == RSS):
                continue;
            else:
                vals.append(dictth[index][key])

        if isdiff == 1 and len(vals) != 0:
            #val_l1.append(vals[0])
            val_l1.extend([j-k for k, j in zip(vals[:-1:2], vals[1::2])])
            new_keys[i]=keys[i][::2]
        else:
            val_l1 = vals
            new_keys[i]=keys[i]

        all_vals[i] = val_l1
        curr_max = 0
        if len(vals) != 0:
            curr_max = max(val_l1)
        if curr_max > max_val:
            max_val = curr_max
        print(len(vals))
        i += 1
        
    return all_vals, new_keys, max_val


def self_aggr(nprocs, lst):
    listf = []
    try:
        print("SELF ", len(lst[0]))
        for i in range(nprocs):
            #print("SELF ", i, " : ", len(lst[i]) )
            if len(lst[i]) != (max_duration*60):
                print("SELF ERR Len : ", len(lst[i]))
                #print("SELF ERR : ", lst[i])
        for j in range(len(lst[0])):
            aggr=0
            val=0
            for i in range(nprocs):
                if len(lst[i]) != 0 and len(lst[i]) >= j:
                    val=lst[i][j]
                    aggr += val 
            listf.append(aggr)
    except:
        print('Index out of range\n')
    return listf

def addition_met(METS, list1, list2):
    listf = {}
    print(METS)
    for i in METS:
        if i <= CMPI:
            l1 = list1[counter_name[i]]
            l2 = list2[counter_name[i]]
            l = [x + y for x,y in zip(l1[:], l2[:])]
            listf[counter_name[i]] = l
    return listf

def addition(nprocs, list1, list2):
    listf = {}
    for i in range(nprocs):
        l1 = list1[i]
        l2 = list2[i]
        l = [x + y for x,y in zip(l1[:], l2[:])]
        listf[i] = l
    return listf

def addition1(nprocs, list1):
    list_tot = {}
    for i in range(nprocs):
        l1 = list1[i]
        tot = 0
        for j in range(len(l1)):
            tot += l1[j]
        list_tot[i] = tot
    return list_tot

def remove_neg(nprocs, negs, val_l1, fixvals=1):
    if negs == None:
        return val_l1

    for i in range(nprocs):
        for neg in negs[i]:
            if neg != 0:
                nval = val_l1[i][neg]
                if neg+1 < len(val_l1[i]) and fixvals == 1: 
                    print("NEG VAL FOR i=", i ," UPDATE PREV: "+ str(val_l1[i][neg+1]))
                    val_l1[i][neg+1] += nval
                    print("NEG VAL FOR i=", i ," UPDATE NOW: "+ str(val_l1[i][neg+1]))
                else:
                    print("NEG VAL  FOR i=", i ," UPDATE PREV: "+ str(val_l1[i][neg+1]))
                del val_l1[i][neg]
    return val_l1

def get_negatives(nprocs, val_l1):
    final_list={}
    for i in range(nprocs):
        negs=[]
        for v in range(len(val_l1[i])):
            if val_l1[i][v] < 0:
                negs.append(v - 1)
                print("NEG FOUND at" + str(v) + " Val=" + str(val_l1[i][v]))
                print(val_l1[i])
        final_list[i] = negs
    return final_list


def intersection(nprocs, list1, list2):
    listf = {}
    for i in range(nprocs):
        l1 = list1[i]
        l2 = list2[i]
        l = []
        l = list(set(l1) & set(l2))
        listf[i] = sorted(l)
    return listf

def union(nprocs, list1, list2):
    listf = {}

    if list1 == None or len(list1) == 0:
        return list2

    if list2 == None or len(list2) == 0:
        return list1

    for i in range(nprocs):
        l1 = []
        l2 =[]
        if len(list1[i]) != 0:
            l1 = list1[i]
        if len(list2[i]) != 0:
            l2 = list2[i]
        l = []
        l = list(set(l1).union(set(l2)))
        listf[i] = sorted(l)

    return listf

def getvalues_sync(dict, nths, val_keys, val_index, isdiff=0):
    i = 0
    max_val = 0
    all_vals = {}
    while i < nths:
        vals = []
        val_l1=[]
        dictth = dict[i]
        keys = sorted(val_keys[i])
        flag = 0
        
        for key in keys:
            if val_index == VHWM:
                print("val ", dictth[val_index][key])
            if dictth[val_index][key] == 0 and (val_index == VHWM or val_index == RSS):
                flag= flag + 1
                if flag == 2:
                   print("val added", dictth[val_index][key])
                   vals.append(dictth[val_index][key])
                   flag = -1
                continue
            else:
                flag = -1                
                print("val added", dictth[val_index][key])
                vals.append(dictth[val_index][key])
    
        if isdiff == 1:
            #vals.sort()
            val_l1.append(vals[0])
            val_l1.extend([j-k for k, j in zip(vals[:-1], vals[1:])])
        else:
            val_l1 = vals
        if val_index == VHWM:
            print("VALS: len", len(val_l1), "val", val_l1 )
        all_vals[i] = val_l1
        curr_max = max(val_l1)
        if curr_max > max_val:
            max_val = curr_max
        i += 1
    return all_vals, max_val

    
def getvalues(dict, nths, index, isdiff=0):
    i = 0
    max_val = 0
    all_vals = {}
    while i < nths:
        vals = [] 
        val_l1=[]
        dictth = dict[i]
        flag = 0
        keys = sorted(dict[i][index])
        for key in keys:
            if dictth[index][key] == 0 and (index == VHWM or index == RSS):
                flag+=1
                if flag == 2:
                   vals.append(dictth[index][key])
                   flag = -1
                continue

                continue;
            else:    
                flag = -1
                vals.append(dictth[index][key])

        if isdiff == 1:
            val_l1.append(vals[0])
            val_l1.extend([j-k for k, j in zip(vals[:-1], vals[1:])])
        else:
            val_l1 = vals

        all_vals[i] = val_l1
        curr_max = max(val_l1)   
        if curr_max > max_val:
            max_val = curr_max
        i += 1
    return all_vals, max_val



def plot_scatter(nprocs, plt, nrows, vals_x, vals_y, title, xlabel, ylabel, max_x, max_y, poincare = 0):
    nrow = int(nrows)
    ncols = 3
    nr = 2 
    k = 1
    fig, sp = plt.subplots(nr, ncols) #, sharex=True, sharey=True)
    #print(int(nrows))
    plt.subplots_adjust(wspace=0.2, hspace=0.26)
    fig.suptitle(title + " Part" + str(k), fontsize=10)
    row_cnt = 0
    j = 0

    for i in range(nprocs):
        labl = "Process " + str(i)

        if j == ncols:
            j = 0
            row_cnt = row_cnt + 1

        if row_cnt == nr:
            row_cnt = 0
            save_plot(plt, title + " Part" + str(k))
            k += 1
            fig, sp = plt.subplots(nr, ncols) #, sharex=True, sharey=True)
            fig.suptitle(title + " Part" + str(k), fontsize=10)
        #print(str(row_cnt) + ',' + str(j))

        color=Colors[i]

        #sp[row_cnt, j].set_ylim([0, max_y]);
        #sp[row_cnt, j].set_xlim([0, max_x]);
        sp[row_cnt, j].set_title(labl, fontsize=8, position=(0.5, 0.8))
        if row_cnt == nr -1:
           sp[row_cnt, j].set_xlabel(xlabel, fontsize=5)
        if j == 0:
           sp[row_cnt, j].set_ylabel(ylabel, fontsize=5)
        if poincare == 1:
            sp[row_cnt, j].scatter(vals_x[i][:-1], vals_y[i][1:], s=1, label=labl, color=color)
        else:
            adj = len(vals_y[i])
            if len(vals_x[i]) < len(vals_y[i]):
                adj = len(vals_x[i])
            elif len(vals_x[i]) > len(vals_y[i]):
                adj = len(vals_y[i])
            sp[row_cnt, j].scatter(vals_x[i][:adj], vals_y[i][:adj], s=1, label=labl, color=color)
 
        j = j + 1

    i += 1
    while i % (ncols * nr) != 0:
        if j == ncols:
            j = 0
            row_cnt = row_cnt + 1
        #print(str(row_cnt) + ',' +str(j))
        sp[row_cnt, j].axis('off')
        j += 1
        i += 1

    save_plot(plt, title + " Part" + str(k))

def timestamp_to_date(date):
    date_ms = [ j/1000000 for j in date] 
    dateconv = numpy.vectorize(dt.datetime.fromtimestamp)
    date1 = dateconv(date_ms)
    return date1.tolist()

def timestamp_to_date1(date):
    dateconv = numpy.vectorize(pd.to_datetime)
    date1 = dateconv(date)
    return date1.tolist()

def timestamp_to_datetime(date):
    #date1 = [Timestamp.to_pydatetime(x) for x in date]
    date1 = [Timestamp.timestamp(x) for x in date]
    return timestamp_to_date(date1)

def round(t, freq):
    freq = to_offset(freq)
    return pd.Timestamp((t.value // freq.delta.value) * freq.delta.value)


def aggregate_x_secs(nprocs, vals, keys, func=0):
    global ini_timestamp
    f_vals={}
    f_keys={}
    max_val = 0
    fake_time = timestamp_to_date([18000000000])
    start_t = timestamp_to_date([ini_timestamp])
    for i in range(nprocs):
        print("Process:", i)
        print("Initial timestamp:", start_t[0])
        val_keys1 = timestamp_to_date(keys[i])
        val_keys = [ x - start_t[0] for x in val_keys1 ]
        val_key = [ fake_time[0] + x  for x in val_keys ]

        print("SIZES OF CONFLICT ->", len(keys[i]), len(vals[i]))
        print("Last timestamp:", val_key[-1])
        if len(keys[i]) > len(vals[i]):        
            val_key=val_key[1:]
            #print(vals[i])
        f_vali = []
        f_keyi = []
        first= val_key[0]
        prev= vals[i][0]
        last = val_key[-1]
        x = 0
        k = 0
        f_keyi.append(first)
        f_vali.append(prev)
        if is_microseconds == 0:
            first = first + dt.timedelta(seconds=frequency)
        else:
            first = first + dt.timedelta(microseconds=frequency)

        while first <=last:
            print("Cuurent timestamp", first, " " , val_key[x])
            prev=vals[i][x]
            while first >= val_key[x] and x < len(vals[i]):
                if func == 0:
                    prev= prev + vals[i][x]
                else:
                    prev= vals[i][x] 
                x = x + 1
            #if diff == 1:
            #    f_vali.append(prev - f_vali[k])
            #else:
            f_vali.append(prev)
            f_keyi.append(first)  
            k = k + 1
            if is_microseconds == 0:
                first = first + dt.timedelta(seconds=frequency)
            else:
                first = first + dt.timedelta(microseconds=frequency)

        dataframe = pd.Series(data=f_vali[:], index=f_keyi[:])
        #dataframe = dataframe1.tz_localize('UTC')

        #print(dataframe)
        #grouped = dataframe.groupby(dataframe.index.map(lambda t: t.minutes)).aggregate(numpy.sum)
        #grouped = dataframe.groupby(pd.Grouper(freq=sample_freq)).aggregate(numpy.sum)
        #grouped1=None
         
        #if func == 0:
        #    grouped1 = dataframe.groupby(partial(round, freq=sample_freq)).sum()
        #else:
        #    grouped1 = dataframe.groupby(partial(round, freq=sample_freq)).last()
           
        t1 = fake_time[0] + dt.timedelta(minutes=max_duration)
        print("TRUNCATE",t1)
        grouped = dataframe.truncate(before='1970-01-01', after='1970-01-02')
        f_vali=grouped.tolist()
        f_keyi=list(grouped.keys())

        first = f_keyi[0]
        last = f_keyi[-1]
        f_keyx = []
        f_valx = []
        x = 0

        while first <= last:
           if first != f_keyi[x]:
               f_keyx.append(first)
               f_valx.append(0)
           else:
               f_keyx.append(f_keyi[x])
               #if diff == 1 and x != 0:
               #    f_valx.append(f_vali[x]- f_vali[x-1])
               #else:
               f_valx.append(f_vali[x])
               x += 1  
           if is_microsecond == 0:  
               first = first + dt.timedelta(seconds=frequency)
           else:
               first = first + dt.timedelta(microseconds=frequency)

        f_keyi = f_keyx
        f_vali = f_valx

        print(grouped)
        #print(f_keyi)
        print("KEY_LEN ", len(f_keyi), len(f_vali)) 

        t2 = fake_time[0]

        time_add = []
        val_add = []

        while f_keyi[0] > t2:
           time_add.append(t2)
           val_add.append(0)
           if is_microsecond == 0:
               t2 = t2 + dt.timedelta(seconds=frequency)
           else: 
               t2 = t2 + dt.timedelta(microseconds=frequency) 
        
        print("PREPEND ", f_keyi[0] , " ", time_add)            
        print("PREPEND ", val_add)
            
        if len(time_add) != 0:
            val_add.extend(f_vali)
            time_add.extend(f_keyi)
            f_keyi = time_add
            f_vali = val_add


        while f_keyi[-1] < t1:
            last_t = f_keyi[-1]
            if is_microsecond == 0:
                last_t = last_t + dt.timedelta(seconds=frequency)
            else:
                last_t = last_t + dt.timedelta(microseconds=frequency)
            f_keyi.append(last_t)
            f_vali.append(0)
        #print(f_keyi)
        #print(f_vali)
 
        f_vals[i]=f_vali
        f_keys[i]=f_keyi

        #print(f_keys[i])
        #print(f_vals[i])
        max_i = max(f_vals[i])
        if max_i > max_val:
           max_val = max_i
        dataframe={}
    #exit(0)
    return f_vals, f_keys, max_val

def plot_timeseries(nprocs, plt, nrows, val_keys, vals, title, xlabel, ylabel,  max_y, path, rAvg=0):
    global ini_timestamp
    nrow = int(nrows)
    ncols = 3
    nr = 2
    k = 1
    fig, sp = plt.subplots(nr, ncols) #, sharex=True, sharey=True)
    #print(int(nrows))
    plt.subplots_adjust(wspace=0.2, hspace=0.26)
    fig.suptitle(title  + " Part" + str(k), fontsize=10)
    row_cnt = 0
    j = 0

    for i in range(nprocs):
        labl = "Process " + str(i)
        print(labl)
        if j == ncols:
            j = 0
            row_cnt = row_cnt + 1
        if row_cnt == nr:
            row_cnt = 0
            save_plot(plt, title + " Part" + str(k))
            k += 1
            fig, sp = plt.subplots(nr, ncols) #, sharex=True, sharey=True)
            fig.suptitle(title + " Part" + str(k), fontsize=10)

        #print(str(row_cnt) + ',' + str(j))

        color=Colors[i]
        datacols = {}
        cols = []

        if len(val_keys[i]) == 0:
            j = j + 1
            continue
        fake_time = timestamp_to_date1([18000])
        #print("FAKE TIME",fake_time)
        val_keyi = timestamp_to_date(val_keys[i])
        val_key1 = [ x - start_t[0] for x in val_keyi ]
        val_key = [ fake_time[0] + x  for x in val_key1 ]
        #val_key = val_keys[i]
        val_key=timestamp_to_datetime(val_keys[i])

        print("COUNT", val_key)
        adj = min([len(val_key), len(vals[i])])
        lend = False
        if rAvg == 1:
             rmean = running_mean(vals[i])
             rmean5=[0,0,0,0]
             rmean5.extend(running_mean2(vals[i], 5))
             datacols={ "Series":vals[i][:adj], "Rmean":rmean[:adj], "Rmean5":rmean5[:adj]}
             cols = ["Series", "Rmean", "Rmean5"]
             #cols = ["OriginaRS']
             sty = ['-o', '-o', '-o' ]
             lend = True
        if rAvg == 2:
             rmin = running_min(vals[i])
             rmax = running_max(vals[i])
             #rstdev, rvariance = running_stdev(vals[i])
             #variance = running_dev(vals[i])
             datacols={ "Series":vals[i][:adj], "Rmin":rmin[:adj], "Rmax":rmax[:adj]} #, "RStd":rstdev[:adj], "RVar":rvariance[:adj] }
             cols = ["Series", "Rmin", "Rmax"] #, "RStd", "RVar"]
             sty = ['-o', '-o', '-o'] #, '-o', '-o' ]

        else:
             datacols={"Series":vals[i][:adj]}
             cols = ["Series"]
             sty = ['-d' ]
             #cols = ['O']

        dataframe = pd.DataFrame(data=datacols, index=val_key[:adj], columns=cols)
        #dataframe = pd.DataFrame(data=datacols, columns=cols)
        #print(dataframe)
        path1=image_path + "/" + path 
        dataframe.to_csv(path_or_buf=path1, sep=' ', index=False, header=True)
        sp[row_cnt, j].set_ylim([0, max_y+1]);
        sp[row_cnt, j].set_xlim([fake_time[0], fake_time[0] + dt.timedelta(minutes=max_duration)]);
        sp[row_cnt, j].set_xticks(numpy.arange(fake_time[0], fake_time[0] + dt.timedelta(minutes=max_duration), dt.timedelta(minutes=2)))
        ax = dataframe.plot(ax=sp[row_cnt, j], legend=lend, ms=1, rot=45, fontsize=6)
        #plt.plot(x=val_key[:adj], y=vals[i][:adj], legend=lend, ms=1, rot=45, fontsize=6)
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%M:%S'))

        #ax.xaxis.set_major_locator(mdates.SecondLocator(interval=int(len(val_key)/10)))
        sp[row_cnt, j].set_title(labl, fontsize=8, position=(0.5, 0.8))
        if row_cnt == nr - 1:
           sp[row_cnt, j].set_xlabel(xlabel, fontsize=7)
        if j == 0 :
           sp[row_cnt, j].set_ylabel(ylabel, fontsize=7)
        j = j + 1
        
    i += 1
    while i % (ncols * nr) != 0:
        if j == ncols:
            j = 0
            row_cnt = row_cnt + 1
        #print(str(i) + ':' + str(row_cnt) + ',' +str(j))
        sp[row_cnt, j].axis('off')
        j += 1
        i += 1
    #plt.show()
    save_plot(plt, title  + " Part" + str(k))


def divideby(nprocs, val_l, scale):
    result = {}
    max_val = 0
    for i in range(nprocs):
        result[i]=[j/scale for j in val_l[i]]
        if max_val < max(result[i]):
             max_val = max(result[i])  
    return result, max_val

def map_vals(nprocs, dict, dict_indx, key1, key2):
    mapped_vals = {}
    max_val = 0
    for i in range(nprocs):
        new_vals = []
        prev_index = -1
        value = 0
        #ar = numpy.asarray(sorted(key2[i]))
        ar = sorted(key2[i])
        #print(key2[i])
        #print(ar) 
        val_l = dict[i]
        for key in key1[i]:
            mval = [x for x in ar if x < key]
            index = len(mval) - 1 
            print("Index is", index, " Previous index is,", prev_index) 
            print("Key is", key, " , Key at index is", mval[index])
            #index = indexList[0]
            if prev_index != -1:
                value1 = val_l[dict_indx][mval[index]]
                value2 = val_l[dict_indx][mval[prev_index]]
                value = value1 - value2
            else:
                value = val_l[dict_indx][mval[index]]
            prev_index = index
            #prev_index = -1 
            new_vals.append(value)
        if len(new_vals) != 0 and max_val <= max(new_vals):
            max_vals = max(new_vals)
        mapped_vals[i] = new_vals
        #print("Maaped to rss", i, "   ", mapped_vals[i])
    return mapped_vals, max_vals

def all_less_100(ls):
    lst=[]
    for i in range(len(ls)):
        if ls[i] > 100:
            lst.append(i)
    return lst

def calculate_percentage1(nprocs, val_l, val_tot, keys):
    percentage = {}
    max_p = 0
    for i in range(nprocs):
        percentage[i]=[ x/y if y else 0 for x,y in zip(val_l[i] , val_tot[i])]
        if len(percentage[i]) != 0 :
            #print("...percentage Percentage..")
            #print(percentage[i])
            if max(percentage[i]) == 0:
                 print("Error:")
                 #print(percentage[i])
                 #exit
        if len(percentage[i]) != 0 and max_p < max(percentage[i]):
             max_p = max(percentage[i])
    return percentage, max_p


def calculate_ratio(nprocs, val_l, val_tot, keys):
    ratio = {}
    max_p = 0
    for i in range(nprocs):
        print(val_l[i] , val_tot[i])
        ratio[i]=[ x/y if y else 0 for x,y in zip(val_l[i] , val_tot[i])]
        if len(ratio[i]) != 0 and max_p < max(ratio[i]):
             max_p = max(ratio[i])
    return ratio, max_p

def calculate_bandwidth(nprocs, val_l, val_tcyc):
    bandwidth = {}
    for i in range(nprocs):
        bandwidth[i]=[(((x * 64)/y) * 1.333) if y else 0 for x,y in zip(val_l[i] , val_tcyc[i])]
    return bandwidth

def running_mean(x):
    cumsum = numpy.cumsum(numpy.insert(x, 0, 0)) 

def calculate_percentage(nprocs, val_l, val_tot, keys):
    percentage = {}
    max_p = 0
    for i in range(nprocs):
        print(val_l[i] , val_tot[i])
        percentage[i]=[ x/y*100 if y else 0 for x,y in zip(val_l[i] , val_tot[i])]
        if len(percentage[i]) != 0 :
            #print("...percentage Percentage..")
            #print(percentage[i])
            err = all_less_100(percentage[i])
            if len(err) != 0:
                 for e in err:
                     print("Error at position:", e, " for process : ", i, "and ", val_l[i][e] , val_tot[i][e], keys[i][e])
                     percentage[i][e] = 0
            if max(percentage[i]) == 0:
                 print("Error:")
                 #print(percentage[i])
                 #exit
        if len(percentage[i]) != 0 and max_p < max(percentage[i]):
             max_p = max(percentage[i])
    return percentage, max_p

def calculate_bandwidth(nprocs, val_l, val_tcyc):
    bandwidth = {}
    for i in range(nprocs):
        bandwidth[i]=[(((x * 64)/y) * 1.333) if y else 0 for x,y in zip(val_l[i] , val_tcyc[i])]
    return bandwidth

def running_mean(x):
    cumsum = numpy.cumsum(numpy.insert(x, 0, 0)) 
    return cumsum

def calculate_percentage(nprocs, val_l, val_tot, keys):
    percentage = {}
    max_p = 0
    for i in range(nprocs):
        print(val_l[i] , val_tot[i])
        percentage[i]=[ x/y*100 if y else 0 for x,y in zip(val_l[i] , val_tot[i])]
        if len(percentage[i]) != 0 :
            #print("...percentage Percentage..")
            #print(percentage[i])
            err = all_less_100(percentage[i])
            if len(err) != 0:
                 for e in err:
                     print("Error at position:", e, " for process : ", i, "and ", val_l[i][e] , val_tot[i][e], keys[i][e])
                     percentage[i][e] = 0
            if max(percentage[i]) == 0:
                 print("Error:")
                 #print(percentage[i])
                 #exit
        if len(percentage[i]) != 0 and max_p < max(percentage[i]):
             max_p = max(percentage[i])
    return percentage, max_p

def calculate_bandwidth(nprocs, val_l, val_tcyc):
    bandwidth = {}
    for i in range(nprocs):
        bandwidth[i]=[(((x * 64)/y) * 1.333) if y else 0 for x,y in zip(val_l[i] , val_tcyc[i])]
    return bandwidth

def running_mean(x):
    cumsum = numpy.cumsum(numpy.insert(x, 0, 0)) 
    sz = cumsum.size
    N = list(range(1,sz))
    #print(cumsum)
    #print(N)
    avg = [x/y for x,y in zip(cumsum, N)]
    return avg

def running_mean2(x, N):
    cumsum = numpy.cumsum(numpy.insert(x, 0, 0)) 
    return (cumsum[N:] - cumsum[:-N]) / float(N)

def running_min(x):
    min_x=[]
    for i in range(len(x)):
        if i == 0:
            min_x.append(x[0])
        else:
            min_x.append(min(x[:i]))
    return min_x

def running_max(x):
    max_x=[]
    for i in range(len(x)):
        if i == 0:
            max_x.append(x[0])
        else:
            max_x.append(max(x[:i]))
    return max_x

def running_stdev(x):
    std_x=[]
    var_x=[]
    for i in range(len(x)):
        #print("ITERATION ", i)
        if i <= 1:
            std_x.append(0)
            var_x.append(0)
        else:
            std_x.append(statistics.stdev(x[:i]))
            var_x.append(statistics.variance(x[:i]))
    return std_x, var_x


def save_plot(plt, title):
    global image_path
    plt.savefig(image_path + "/" + title + ".png")

def merge_keys(key1, key2, nprocs): 
    merged_key={}
    for i in range(nprocs): 
        if key1.get(i) != None:
            #print(key1[i])
            #print(key2[i])
            merged_key[i] = key1[i]
            merged_key[i].extend(key2[i])
            merged_key[i] = sorted(merged_key[i])
            #print(merged_key[i])
        elif key2.get(i) != None:
            merged_key[i] = key2[i]
        else:
            merged_key[i] = []
    return merged_key


def main():
    global ini_timestamp
    global image_path
    if len(sys.argv) != 8:
       print("Usage: python tau_python.py tracedir nprocs nthreads image_path system")
       print(sys.argv)
    exp_dir = sys.argv[1]
    nprocs1 = int(sys.argv[2])
    nprocs2 = int(sys.argv[3])
    nprocs3 = int(sys.argv[4])
    ini_timestamp = int(sys.argv[5])
    image_path= sys.argv[6]
    system = sys.argv[7]
    reference = 0
    max_mem = max_papi = max_per = 0
    dirs = os.listdir(exp_dir)
    vals_final = None    
    #Mets=['RSS', 'L3', 'TCA', 'MEMO', 'INS', 'TCYC', 'LD', 'ST', 'VmSwap']
    #Mets=['RSS', 'MEMO', 'INS', 'TCYC', 'LD', 'ST', 'VmSwap']
    Mets = None
    str_p = 0
    keys_papi = None
    print(dirs)
    for dir in dirs:
        print("Processing  " + str(dir))
        if os.path.isdir(exp_dir + "/" + dir) == False:
            continue
        vals_dirs = None
        max_memt = max_papit = max_pert = 0
        tracedir = exp_dir + "/" + dir
        if dir == "simulation":
            reference = 0
            edffile = tracedir + "/events." + str(reference) + ".edf"
            #outpath= image_path + "/Simulation"
            outpath= image_path + "/" + dir
            vals_dir, keys_mem, keys_papi, max_memt, max_papit, max_pert,counter = get_max(outpath, edffile, tracedir, nprocs1, 1, reference)
        elif bool("simulation" in dir):
            reference = int(dir.split('_')[1])
            np = int(dir.split('_')[2])
            edffile = tracedir + "/events." + str(reference) + ".edf"
            #outpath= image_path + "/Simulation"
            outpath= image_path + "/" + dir
            vals_dir, keys_mem, keys_papi, max_memt, max_papit, max_pert,counter = get_max(outpath, edffile, tracedir, np, 1, reference, system)

        if dir == "subsample":
            #reference = nprocs1
            reference = 0 #nprocs1
            edffile = tracedir + "/events." + str(reference) + ".edf"
            #outpath= image_path + "/Particle_selection"
            outpath= image_path + "/" + dir
            vals_dir, keys_mem, keys_papi, max_memt, max_papit, max_pert,counter = get_max(outpath, edffile, tracedir, nprocs2, 1, reference, system)
        elif bool("subsample" in dir):
            reference = int(dir.split('_')[1])
            np = int(dir.split('_')[2])
            print("reference ", reference)
            edffile = tracedir + "/events." + str(reference) + ".edf"
            outpath= image_path + "/" + dir
            vals_dir, keys_mem, keys_papi, max_memt, max_papit, max_pert,counter = get_max(outpath, edffile, tracedir, np, 1, reference, system)
        elif dir == "particle" or bool("particle" in dir) or bool('analysis' in dir):
            np = None  
            if dir == 'particle':
                #reference = nprocs2 + nprocs1
                reference = 0 #nprocs2 + nprocs1
                print("particle:reference", reference, " nprocs2 ", nprocs2, " nprocs1 " , nprocs1 )
                np = nprocs3
            elif bool("particle" in dir):
                reference = int(dir.split('_')[1])
                np = int(dir.split('_')[2])
                edffile = tracedir + "/events." + str(reference) + ".edf"
                outpath= image_path + "/" + dir
                vals_dir, keys_mem, keys_papi, max_memt, max_papit, max_pert,counter = get_max(outpath, edffile, tracedir, np, 1, reference, system)
            elif dir == 'analysis':
                #reference = nprocs1
                reference = 0
                np = nprocs2
            elif dir == 'analysis1':
                reference = nprocs2 + nprocs1
                np = nprocs3
            elif dir == 'analysis2':
                reference = nprocs3 + nprocs2 + nprocs1
                np = nprocs3
            outpath= image_path + "/" + dir
            edffile = tracedir + "/events." + str(reference) + ".edf"
            vals_dir, keys_mem, keys_papi, max_memt, max_papit, max_pert, counter = get_max(outpath, edffile, tracedir, np, 1, reference, system)
        elif bool('stream' in dir):
            reference = 0
            edffile = tracedir + "/events." + str(reference) + ".edf"
            outpath= image_path + "/" + dir
            vals_dir, keys_mem, keys_papi, max_memt, max_papit, max_pert,counter = get_max(outpath, edffile, tracedir, 1, 1, 0, system)
        else:
            continue
        print("\nProcessed", tracedir) 
        if Mets == None:
            Mets = counter
        if vals_final == None:
            vals_final = vals_dir
        else:
            vals_final = addition_met(Mets, vals_final, vals_dir)
        if (max_mem > max_memt):
            max_mem = max_memt
        if (max_papi > max_papit):
            max_papi3 = max_papt
        if (max_per > max_pert):
            max_per = max_pert

    max_memr = max_papir = max_perr = 0
    statdir = exp_dir + "/../stat.txt"
    print("STAT FILE:", statdir)
    if os.path.isfile(statdir) == True:
        max_list = []
        with open(statdir, 'r') as f:
            print("STAT FILE OPENED:", statdir)
            maxstr = f.readline()
            max_list = maxstr.split()
            print("MAX VALUES READ: ", max_list)
        max_memr = float(max_list[0])
        max_papir = max_l2_abs = max_l3_abs = int(max_list[1])
        max_perr = max_p_l2 = max_p_l3 = float(max_list[2])
    statdir = exp_dir + "/../stat.txt"
    print("STAT FILE:", statdir)
    with open(statdir, 'w') as f:
        f.truncate()
        max_mem = max([max_mem, max_memr])
        max_papi = max([max_papir, max_papi])
        max_per = max([max_perr, max_per])
        f.write(str(max_mem)+ " " + str(max_papi) + " " + str(max_per))
        print("WRITE" + str(max_mem)+ " " + str(max_papi) + " " + str(max_per))
   ## ----------------------- Write the statistics to a file --------------------------------------------------- ##
    xlsx_fname = "overall-stats.xlsx"
    save_data(image_path, xlsx_fname, "Overall", vals_final, keys_papi, 1)


def get_max(expdir, edffile, tracedir, nprocs, nthreads, reference, system):
    global function_id
    global tag
    global unmap
    global mylist
    global umpi

    function_id = {}
    tag = {}
    unmap = {}
    umpi = {}
    mylist = {}
    keys = {}
    vals= {}
    umpi[MPIS%CMPI] = []
    umpi[MPIR%CMPI] = []
    umpi[MPISR%CMPI] = []

    for ctr in counter_name.keys():
        if ctr != MPIS and ctr != MPISR and ctr != MPIR:
            unmap[ctr] = -1
            keys[ctr]={}
            vals[ctr]={}
        else:
            umpi[ctr%CMPI] = []

    counter = read_edf(edffile, system)
    read_trace(tracedir, nprocs, nthreads, reference)
    
    print("Reading keys of counter RSS")
    keys[RSS] = getAllKeys(mylist, nprocs, RSS)
    print("Reading keys of counter VHWM")
    keys[VHWM] = getAllKeys(mylist, nprocs, VHWM)
    print("KEYS: ", len(keys[VHWM][0])," ", len(keys[RSS][0]))
    for ctr in counter:
        if ctr != RSS and ctr != VHWM and ctr != MPIS and ctr != MPIR and ctr != MPISR:
            print("Reading keys of counter " + counter_name[ctr])
            keys[ctr] = getAllKeys(mylist, nprocs, ctr)

    keys_papi = None
    for ctr in counter:
        if ctr == RSS or ctr == VHWM or ctr == MPIS or ctr == MPIR or ctr == MPISR:
           continue; 
        print("Finding common keys " + counter_name[ctr])
        if keys_papi == None:
           keys_papi = keys[ctr]
        else:
           keys_papi = intersection(nprocs, keys_papi, keys[ctr])

    max_papi = 0
    for ctr in counter:
       print("Reading values of " + counter_name[ctr])
       if ctr == RSS or ctr == VHWM: 
          vals[ctr], max_mem = getvalues_sync(mylist, nprocs, keys[ctr], ctr)
       elif ctr != MPIS and ctr != MPISR and ctr != MPIR:
          vals[ctr], max_val = getvalues_sync(mylist, nprocs, keys_papi, ctr, 1)
          if ctr == L3:
             max_papi = max_val
       else:
            vals[ctr], max_val = getvalues_sync(mylist, nprocs, keys[ctr], ctr, 1)


    negs = None 
    for ctr in counter:
       if ctr != RSS and ctr != VHWM and ctr != MPIS and ctr != MPISR and ctr != MPIR:
          neg = get_negatives(nprocs, vals[ctr])
          if negs != None:
             negs = union(nprocs, negs, neg)
          else:
             negs = neg

    for ctr in counter:
       if ctr != RSS and ctr != VHWM and ctr != MPIS and ctr != MPISR and ctr != MPIR:
          vals[ctr] = remove_neg(nprocs, negs, vals[ctr])

    keys_papi = remove_neg(nprocs, negs, keys_papi, 0)

    for ctr in counter:
        if ctr == RSS or ctr == VHWM:
            print("C1: formatting values of " + counter_name[ctr])
            vals[ctr] = aggregate_x_secs(nprocs, vals[ctr], keys[ctr], 1)
        elif ctr != MPIS and ctr != MPISR and ctr != MPIR:
            print("C2: formatting values of " + counter_name[ctr])
            vals[ctr] = aggregate_x_secs(nprocs, vals[ctr], keys_papi)

    scale=1000000
    vals[RSS], max_mem = divideby(nprocs, vals[RSS][0], scale)
    vals[VHWM], max_mem = divideby(nprocs, vals[VHWM][0], scale)
    max_per = 100
    print("After scaling", vals[RSS])
    vals_final = None
    for i in range(nprocs):
        datadir = "data_process_" + str(i) + ".xlsx"
        sheet = "Process_" + str(i)
        vals_final={}
        for ctr in counter:
           if ctr == MPISR or ctr == MPIS or ctr == MPIR:
               continue
           if ctr != RSS and ctr !=VHWM:
               vals_final[counter_name[ctr]] = vals[ctr][0][i]
           else:
               vals_final[counter_name[ctr]] = vals[ctr][i]
        save_data(expdir, datadir, sheet, vals_final, keys_papi, nprocs)      
    vals_all={}
    for ctr in counter:    
        print("adding values of " + counter_name[ctr])
        if ctr == MPISR or ctr == MPIS or ctr == MPIR:
           continue
        if  ctr == RSS or ctr == VHWM: 
           vals_all[counter_name[ctr]] = self_aggr(nprocs, vals[ctr])
        elif ctr != MEM1_RD1 and ctr != MEM1_RD2 and ctr != MEM1_RD3 and ctr != MEM1_RD4 and ctr != MEM1_WR1 and ctr != MEM1_WR2 and ctr != MEM1_WR3 and ctr != MEM1_WR4 and ctr != MEM2_RD1 and ctr != MEM2_RD2 and ctr != MEM2_RD3 and ctr != MEM2_RD4 and ctr != MEM2_WR1 and ctr != MEM2_WR2 and ctr != MEM2_WR3 and ctr != MEM2_WR4 and ctr != RQ_RD1 and ctr != RQ_RD2 and ctr != RQ_WR1 and ctr != RQ_WR2 and ctr != IR_RD1 and ctr != IR_RD2 and ctr != IR_WR1 and ctr != IR_WR2 and ctr != DR_A and ctr != DR_O and ctr != SQ_F: 
           val = vals[ctr][0]
           vals_all[counter_name[ctr]] = self_aggr(nprocs, val)
        else:
           vals_all[counter_name[ctr]] = vals[ctr][0][0]
    #print(vals_all)
    xls_fname = "data.xlsx"    
    save_data(expdir, xls_fname, 'Combined', vals_all, keys_papi, nprocs)
    return vals_all, keys[RSS], keys_papi, max_mem, max_papi, max_per,counter

def write_to_excel(worksheet, row, col, lst, title):
    worksheet.write(row-1, col, title)
    for key in range(len(lst)):
        item=lst[key]
        worksheet.write(row, col, item)
        row += 1


def write_to_excel1(fname, sheetname, row, col, lst, title):
    workbook = load_workbook(filename = fname)
    worksheet = workbook[sheetname]
    worksheet.cell(row=row, column=col).value=title
    for key in range(len(lst)):
        row += 1
        if lst[key] != 0: 
            worksheet.cell(row=row, column=col).value = lst[key]
    workbook.save(fname);

def save_data(out_dir, xlsx_fname, sheetname, vals_final, keys_papi, nprocs):

    sheetname = 'Experiment'
    #copy_rename('Sample_format.xlsx', xlsx_fname, '.', out_dir)
    xlsx_fname = out_dir + "/" + xlsx_fname
    create_workbook(xlsx_fname, sheetname)
    #print("VALUEEES ", vals_final)
    row = 1
    col = 3
    nitems = len(vals_final['RSS'])
    app = os.path.basename(out_dir)
    apps_lst = []
    nprocs_lst = []  
    label_lst = []
    for i in range(nitems):
        apps_lst.append(app)
        nprocs_lst.append(nprocs)
        label_lst.append(0)
 
     
    #write_to_excel1(xlsx_fname, sheetname, row, col, apps_lst, "Application")
    #col = col + 1 

    #write_to_excel1(xlsx_fname, sheetname, row, col, nprocs_lst, "Nprocess")
    #col = col + 1 

    #write_to_excel1(xlsx_fname, sheetname, row, col, label_lst, "Label")
    #col = col + 1 

    if vals_final.get('RSS', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['RSS'], "RSS")
        col = col + 1 

    row = 1
    flag = 0
    val_l3 = {}
    val_tca = {}
    val_ld = {}
    val_ip = {}
    val_cyc = {} 
    if vals_final.get('L3', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['L3'], "L3_Misses")
        flag = 1
        val_l3[0] = vals_final['L3']
        col = col + 1 

    if vals_final.get('TCA', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['TCA'], "L3 Cache References")
        flag = flag + 1
        val_tca[0] = vals_final['TCA']
        col = col + 1 

    if flag == 2:
        l3_per2, max_p_l3 = calculate_percentage(1, val_l3, val_tca, keys_papi)
        row = 1
        write_to_excel1(xlsx_fname, sheetname, row, col, l3_per2[0], "L3_Misses_Self_Per")
        col = col + 1 

    flag = 0
    flag2 = 0
    if vals_final.get('LD', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['LD'], "Load_Stalls")
        val_ld[0] = vals_final['LD']
        col = col + 1
        flag = 1 

    if vals_final.get('ST', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['ST'], "Memory Store stalls")
        col = col + 1 
    col = col + 1

    if vals_final.get('VHWM', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['VHWM'], "Vm Swaps")
        col = col + 1

    if vals_final.get('INS', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['INS'], "Instructions retired")
        val_ip[0] = vals_final['INS']
        col = col + 1 
        flag2 = 1
 
    if vals_final.get('TCYC', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['TCYC'], "Unhalted Clock Cycles")
        col = col + 1 
        val_cyc[0] = vals_final['TCYC']
        flag = flag + 1
        flag2 = flag2 + 1

    if flag == 2:
        ld_per2, max_p_l3 = calculate_percentage(1, val_ld, val_cyc, keys_papi)
        row = 1
        write_to_excel1(xlsx_fname, sheetname, row, col, ld_per2[0], "Load_Stalls_Self_Per")
        col = col + 1 

    if flag2 == 2:
        ipc_per2, max_p_l3 = calculate_ratio(1, val_ip, val_cyc, keys_papi)
        row = 1
        write_to_excel1(xlsx_fname, sheetname, row, col, ipc_per2[0], "IPC")
        col = col + 1 

    if vals_final.get('MEMO', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEMO'], "Memory Operations completed")
        col = col + 1 

    if vals_final.get('MEM1_RD1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_RD1'], "Memory reads completed on Socket1  ch 1")
        col = col + 1 

    if vals_final.get('MEM1_WR1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_WR1'], "Memory writes completed on Socket1  ch 1")
        col = col + 1 

    if vals_final.get('MEM1_RD2', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_RD2'], "Memory reads completed on Socket1  ch 2")
        col = col + 1 

    if vals_final.get('MEM1_WR2', None) != None:
       write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_WR2'], "Memory writes completed on Socket1  ch 2")
       col = col + 1 


    if vals_final.get('MEM1_RD3', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_RD3'], "Memory reads completed on Socket1  ch 3")
        col = col + 1 


    if vals_final.get('MEM1_WR3', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_WR3'], "Memory writes completed on Socket1  ch 3")
        col = col + 1 


    if vals_final.get('MEM1_RD4', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_RD4'], "Memory reads completed on Socket1  ch 4")
        col = col + 1 


    if vals_final.get('MEM1_WR4', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_WR4'], "Memory writes completed on Socket1  ch 4")
        col = col + 1 

    if vals_final.get('MEM2_RD1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM2_RD1'], "Memory reads completed on Socket2  ch 1")
        col = col + 1

    if vals_final.get('MEM2_WR1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM2_WR1'], "Memory writes completed on Socket2  ch 1")
        col = col + 1

    if vals_final.get('MEM2_RD2', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM2_RD2'], "Memory reads completed on Socket2  ch 2")
        col = col + 1

    if vals_final.get('MEM2_WR2', None) != None:
       write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM2_WR2'], "Memory writes completed on Socket2  ch 2")
       col = col + 1

    if vals_final.get('MEM2_RD3', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM2_RD3'], "Memory reads completed on Socket2  ch 3")
        col = col + 1
    
    if vals_final.get('MEM2_WR3', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM2_WR3'], "Memory writes completed on Socket2  ch 3")
        col = col + 1

    if vals_final.get('MEM2_RD4', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM2_RD4'], "Memory reads completed on Socket2  ch 4")
        col = col + 1

    if vals_final.get('MEM2_WR4', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['MEM1_WR4'], "Memory writes completed on Socket2  ch 4")
        col = col + 1 

    if vals_final.get('RQ_RD1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['RQ_RD1'], "Total reads issued socket 1")
        col = col + 1 

    if vals_final.get('RQ_WR1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['RQ_WR1'], "Total writes issued socket 1")
        col = col + 1 

    if vals_final.get('RQ_RD2', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['RQ_RD2'], "Total reads issued socket 2")
        col = col + 1 

    if vals_final.get('RQ_WR2', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['RQ_WR2'], "Total writes issued socket 2")
        col = col + 1 

    if vals_final.get('IR_RD1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['IR_RD1'], "Total reads issued socket 1")
        col = col + 1

    if vals_final.get('IR_WR1', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['IR_WR1'], "Total writes issued socket 1")
        col = col + 1

    if vals_final.get('IR_RD2', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['IR_RD2'], "Total reads issued socket 2")
        col = col + 1

    if vals_final.get('IR_WR2', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['IR_WR2'], "Total writes issued socket 2")
        col = col + 1
    
    if vals_final.get('DR_A', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['DR_A'], "All demand read requests")
        col = col + 1

    if vals_final.get('DR_O', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['DR_O'], "Outstanding demand reads")
        col = col + 1

    if vals_final.get('SQ_F', None) != None:
        write_to_excel1(xlsx_fname, sheetname, row, col, vals_final['SQ_F'], "Super Queue Full")
        col = col + 1


if __name__ == "__main__":
    main()	 
