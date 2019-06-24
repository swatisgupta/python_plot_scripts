import sys
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import csv
from mpl_toolkits.mplot3d import Axes3D
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn import metrics
from openpyxl import load_workbook
import xlsxwriter
import seaborn as sns
from filterpy.kalman import KalmanFilter
from filterpy.common import Q_discrete_white_noise

from statistics import median
from statistics import mean
from statistics import variance
from statistics import stdev

from sklearn.externals.six import StringIO  
from IPython.display import Image
from sklearn.tree import export_graphviz
import pydotplus

 
left   =  0.125  # the left side of the subplots of the figure
right  =  0.9    # the right side of the subplots of the figure
bottom =  0.1    # the bottom of the subplots of the figure
top    =  0.9    # the top of the subplots of the figure
wspace =  0.5     # the amount of width reserved for blank space between subplots
hspace =  1.1    # the amount of height reserved for white space between subplots

functions = {'Max': max , 'Avg': mean}
# , 'Std': stdev , 'Mid': median }
all_cols = [ "Application", "L3_Misses", "Load_Stalls", "L3_Misses_Self_Per", "Load_Stalls_Self_Per", "IPC", "Label"]
 
feature_sel = ["L3_Misses_Self_Per", "Load_Stalls_Self_Per", "IPC"]
float_indexes = ["L3_Misses_Self_Per", "Load_Stalls_Self_Per", "IPC"]

sns.set(style="darkgrid")

def load_my_data(csvfile, sp):
    dataset = pd.read_csv(csvfile, sep=sp, encoding='utf-8',  engine='python')
    dataset.fillna(0, inplace=True)
    return dataset

def load_excel_data(xlsfile, sheet):
    dataset = pd.read_excel(xlsfile, sheet_name=sheet)
    dataset.fillna(0, inplace=True)
    return dataset

def get_w_count(list1, app, w, comp, cond = 0):
    i = 0
    count = 0
    while i < w:
        if i != 0 and app[i-1] != app[i]:
            count = 0
        if cond == 0 and list1[i] > comp:
           count = count + 1 
        elif cond == 1 and list1[i] < comp:
           count = count + 1   
        i = i + 1
    return count
         
def get_w_avg(list1, app, w):
    i = 0
    k = 0
    sum = 0
    while i < w:
        if i != 0 and app[i-1] != app[i]:
            sum = 0
            k = 0
        sum = sum + list1[i]
        i = i + 1
        k = k + 1
    return (sum/k)
 
def get_predictions(dataset, app_idx, flist, w):
    res = []
    mr_ap = None
    ipc_ap = None
    ls_ap = None
    print(flist)
    load_stalls = dataset[flist[1]].values.tolist()
    app = dataset[app_idx].values.tolist()
    miss_rate = dataset[flist[0]].values.tolist()
    ipc = dataset[flist[2]].values.tolist()
    total_rows = len(dataset.index)
    print(total_rows)
    for i in range(w):
        res.append(0)
    if total_rows < w:
        return res[0:total_rows]
    k = 1
    for i in range(w, total_rows):
        if app[i] != app[i-1]:
            mr_ap = None
            ipc_ap = None
            ls_ap = None
            k = 1
        ls_y = get_w_count(load_stalls[i-w:], app[i-w:], w, 60)
        ls_a = get_w_avg(load_stalls[i-w:], app[i-w:], w)
        mr_y = get_w_count(miss_rate[i-w:], app[i-w:], w, 60)
        mr_a = get_w_avg(miss_rate[i-w:], app[i-w:], w)
        ipc_y = get_w_count(ipc[i-w:], app[i-w:], w, 1, 1)
        ipc_a = get_w_avg(ipc[i-w:], app[i-w:], w)
        if ipc_ap == None or mr_ap == None or ls_ap == None:
            ipc_ap = ipc_a
            mr_ap = mr_a
            ls_ap = ls_a
        #if (ipc_y > 3 and mr_y > 3 and ls_y > 3) or ((mr_a >= 1.2 * mr_ap) and (ls_a >= 1.2 * ls_ap or ls_a <= 1.2 * ls_ap) and (1.2 * ipc_a <= ipc_ap)):
        if (ipc_y > 3 and ls_y > 3) or ((mr_a >= 1.2 * mr_ap or mr_a <= 1.2 * mr_ap) and (ls_a >= 1.2 * ls_ap) and (1.2 * ipc_a <= ipc_ap)):
            res.append(1)     
        else:
            res.append(0)
        
        ipc_ap = (ipc[i] + (k - 1 ) * ipc_ap)/ k
        mr_ap = (miss_rate[i] + (k - 1) * mr_ap) / k
        ls_ap = (load_stalls[i] + (k - 1) * ls_ap) / k
        k = k + 1 
    return res

def compute_convolution_filter(dataset, label_idx, val_idx, filter):
    vals_o = dataset[val_idx].values.tolist()
    labl_o = dataset[label_idx].values.tolist()
    window = len(filter)
    
    fval = []
    vals = []
    labl = []
    lst = []
    for j in range(window):
        fval.append(0)
    
    vals.extend(fval[1:])
    vals.extend(vals_o)
    vals.extend(fval[1:])
    labl.extend(fval[1:])
    labl.extend(labl_o)
    labl.extend(fval[1:])

    i = window
    l = window
    half_fil = int(window/2)
    while i < len(vals_o) + window:
        if labl[i-1] != labl[i]:
            l = 0  
        else:
            l = l + 1
        if l < half_fil:
            lst.extend([float(0)])
            i = i + 1
            continue   
        sum = 0
        for h in range(window):
            if h < half_fil:
                #if l < half_fil and h < (half_fil - l):
                #    sum = sum + (filter[h] * 0)
                #else: 
                #    sum = sum + ( filter[h] * vals[i - (half_fil - 1 - h)] )
                sum = sum + ( filter[h] * vals[i - (half_fil - 1 - h)] )
            elif h == half_fil:
                sum = sum + filter[h] * vals[i]
            else:
                sum = sum + filter[h] * vals[i + (h - half_fil)]
        lst.extend([float(sum)]) 
        i = i + 1   
    return lst
    
    
def compute_fn_sliding_window(dataset, label_idx, val_idx, window):
    print("Computing Avg, Median, Std for " + val_idx)
    vals = dataset[val_idx].values.tolist()
    labl = dataset[label_idx].values.tolist()
    for fn in functions.keys():
        lst = []
        k = 0
        flag = 0
        for i in range(len(vals)):
            cval = 0
            if i != 0 and labl[i-1] != labl[i]:
                flag = 0
                k = 0
            if flag == 0 and k < window - 1:    
                k = k + 1
            else:
                flag = 1
                cval = functions[fn](vals[i-k:i])
            lst.extend([float(cval)])
        n_lbl = val_idx + "_" + fn + "_" + str(window)
        dataset[n_lbl] = lst
    return dataset

def compute_exp_filter(dataset, label_idx, val_idx):
    vals = dataset[val_idx].values.tolist()
    labl = dataset[label_idx].values.tolist()
    lst = []
    k = 0
    flag = 0
    exp_sum = 0
    alpha = 0.7
    for i in range(len(vals)):
        if i != 0 and labl[i-1] != labl[i]:
             flag = 0
             k = 0
        if flag == 0 and k < window - 1:
             lst.append(vals[i])
             k = k + 1
        else:
             flag = 1
             exp_sum = vals[i] * alpha + (1 - alpha) * exp_sum 
             lst.extend([float(cval)])
        n_lbl = val_idx + "_" + exp_fil
        dataset[n_lbl] = lst
    return dataset

def gaussian_filter(x):  
    y = 1 / np.sqrt(2 * np.pi) * np.exp(-x ** 2 / 2)
    return y
    

def merge_datasets(dataset1, dataset2, indexes):
    df1 = dataset1[indexes]
    df2 = dataset1[indexes]
    frames = [df1, df2]
    result = pd.concat(frames)
    return result
    
def convert_to_numeric(dataset, indexes):
    for index in indexes:
        dataset[index] = dataset[index].astype(float)
        #dataset[index] = dataset[index].astype(int)
    return dataset


def select_features(dataset, sel_f):
    X = dataset[sel_f]
    Y = dataset['Label']
    return X,Y

def verify_correctness(Y_test, Y_pred, dir):
    cnf_matrix = metrics.confusion_matrix(Y_test, Y_pred)
    lst = []
    with open(dir + "/clf_model_results.txt","w+") as f: 
         f.write("Accuracy: " + str(metrics.accuracy_score(Y_test, Y_pred)))
         f.write("\nPrecision: " + str(metrics.precision_score(Y_test, Y_pred)))
         f.write("\nRecall: " + str(metrics.recall_score(Y_test, Y_pred)))
         lst = [metrics.accuracy_score(Y_test, Y_pred), metrics.precision_score(Y_test, Y_pred), metrics.recall_score(Y_test, Y_pred)]
    return cnf_matrix, lst

def plot_conf_matrix(cnf_matrix, axs, fig):
    class_names=[0,1] # name  of classes
    fig, ax = plt.subplots()
    tick_marks = np.arange(len(class_names))
    plt.xticks(tick_marks, class_names)
    plt.yticks(tick_marks, class_names)
    # create heatmap
    p2 = sns.heatmap(pd.DataFrame(cnf_matrix), annot=True, cmap="YlGnBu" ,fmt='g', ax=axs)
    p2.xaxis.set_label_position("top")
    #fig.tight_layout()
    p2.set_title('Confusion matrix', y=1.1)
    p2.set_ylabel('Actual label')
    p2.set_xlabel('Predicted label')
 
def plot_3d(df, idx1, idx2, idx3):
    fig = plt.gcf()
    ax = fig.add_subplot(111, projection='3d')
    #ax.scatter(df[idx1], df[idx2], df[idx3], c='skyblue', s=60)
    fill_colors = ['#FF9999' if wt==1 else '#FFE888' for wt in list(df['Label'])]
    ax.scatter(df[idx1], df[idx2], df[idx3], color=fill_colors)
    ax.set_xlabel(idx1)
    ax.set_ylabel(idx2)
    ax.set_zlabel(idx3)    
    #ax.view_init(30, 185)
    #plt.show()

def plot_3d_2d(df, idx1, idx2, idx3):
    df_3d = pd.DataFrame()
    df_3d[idx1] = df[idx1]
    df_3d[idx2] = df[idx2] 
    #df_3d[idx3] = df[idx3]
    df_3d["Label"] = df["Label"]
    sns.pairplot(df_3d, hue="Label") 

def main():
    
    if len(sys.argv) != 4:
        print("Please provide window size1, window size2 and output directory!", len(sys.argv))
        exit(-1)

    window = int(sys.argv[1])
    window2 = int(sys.argv[2])
    input_dir = sys.argv[3]
    out_dir = sys.argv[3] + "filter_" + str(window) + "_" + str(window2)
   
    #filename = input_dir + "/overall-stats.xlsx"
    #dataset1 = load_excel_data(filename, 'Experiment')
    #dataset1 = preprocess_data(dataset1, window, window)
  
    #process_file("overall-stats", dataset1, window, window2, out_dir)
    
    os.mkdir(out_dir)
    dirs = os.listdir(input_dir) 
    for dir in dirs:
        if os.path.isdir(input_dir + "/" + dir) == False:
            continue 

        filename = input_dir + "/" + dir + "/data.xlsx"        
        dataset1 = load_excel_data(filename, 'Experiment')
        dataset1 = preprocess_data(dataset1, window, window2)
        process_file("data-" + dir, dataset1, window, window2, out_dir)


def process_file(filename, dataset, window, window2, outdir):
    otf = filename + "_" + str(window) + "_" + str(window2) + ".xlsx"  
    d_set_names = [otf]
    d_sets = [dataset]
    #os.mkdir(outdir + "/" + otf)
    print("Processing overall dataset")
    print(dataset.head(10))
    create_experiments_sel(dataset, window, window2, outdir, feature_sel, d_sets, d_set_names) 

     
    

def create_experiments_sel(dataset, window, window2, dir, ft_col, test_in, test_dname):
    str_f = ""
    for i in ft_col:
        str_f = str_f + "_" + i
    #exp_dir = dir+"/w"+ str(window) + str_f
    #os.mkdir(exp_dir)
    f_cols = ft_col
    fl_cols = []
    fl_cols1 = []
    k_cols = []
    #output = get_apply_model(dataset, f_cols, 2, exp_dir, window2)
    #with open(dir + "/Summary.csv", "a+") as s:
    #    s = csv.writer(s, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    #    s.writerow(output)

    for k in functions.keys():
        lbl = ""
        str_f = ""
        for i in ft_col:
            str_f = str_f + "_" + i + "_" + k
            lbl = i + "_" + k + "_" + str(window)
            fl_cols.append(lbl)
            for l in functions.keys():
                if l != k:
                    lb = lbl+ "_" + l + "_" + str(window)
                    fl_cols1.append(lb)
    for i in f_cols:
        lbl = i + "_kalman"
        k_cols.append(lbl) 

    dn = 0
    row = 1
    col = 1     
    cl = col   

    for dset in test_in:
        create_workbook(dir + "/" + test_dname[dn] + ".xlsx", 'test_outcome')
        cl = col
        for i in ['Application', 'Label']:
            lst = dset[i].values.tolist()
            write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , lst, i)
            cl = cl + 1
        for i in f_cols:
            lst = dset[i].values.tolist()
            write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , lst, i)
            cl = cl + 1
        for i in k_cols:
            lst = dset[i].values.tolist()
            write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , lst, i)
            cl = cl + 1  
        #for i in fl_cols:
        #    lst = dset[i].values.tolist()
        #    write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , lst, i)
        #    cl = cl + 1
        for i in fl_cols1:
            lst = dset[i].values.tolist()
            write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , lst, i)
            cl = cl + 1  
         
        dn = dn + 1

    col = cl
    dn = 0
    nfeatures = len(f_cols)

    #for dset in test_in:
    #    cl = col
    #    test_pred_acc(dset, test_dname[dn], dir, row, cl, f_cols, 'raw', window2)
        #dn = dn + 1
        #exp_dir = dir+"/w"+ str(window) + str_f
        #os.mkdir(exp_dir)
        #output = get_apply_model(dataset, f_cols, 2, exp_dir, window2)
        #with open(dir + "/Summary.csv", "a+") as s:
        #    s = csv.writer(s, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        #    s.writerow(output)
    #    dn = 0
    #    cl = cl + 1
    #    nf = 0
    #    for k in functions.keys():
    #        test_pred_acc(dset, test_dname[dn], dir, row, cl, fl_cols[nf:nf+nfeatures], k, window2)
    #        nf = nf + nfeatures   
    #        cl = cl + 1
    #    dn = dn + 1

    dn = 0
    col = cl
    filter_naive = []
    filter_exp = []
    filter_log = []

    for x in range(window2):
        filter_naive.append(-1)
    for x in range(window2):
        filter_naive.append(1)

     

    filter = filter_naive

    for dset in test_in:
        cl = col
        #for x in f_cols:
        #    res = compute_convolution_filter(dataset, 'Application', x, filter)
        #    write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , res, x + "_conv" + "_" + str(2 * window2))
        #    cl = cl + 1
        for x in k_cols:
            res = compute_convolution_filter(dataset, 'Application', x, filter)
            write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , res, x + "_conv" + "_" + str(2 * window2))
            cl = cl + 1
        #for x in fl_cols:
        #    res = compute_convolution_filter(dataset, 'Application', x, filter)
        #    write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , res, x + "_conv" + "_" + str(2 * window2))
        #    cl = cl + 1 
        for x in fl_cols1: 
            res = compute_convolution_filter(dataset, 'Application', x, filter)
            write_to_excel1(dir + "/" + test_dname[dn] + ".xlsx", "test_outcome", row, cl , res, x + "_conv" + "_" + str(2 * window2))
            cl = cl + 1
        dn = dn + 1 


def write_to_excel1(fname, sheetname, row, col, lst, title):
    workbook = load_workbook(filename = fname)
    worksheet = workbook[sheetname]
    worksheet.cell(row=row, column=col).value=title
    for key in range(len(lst)):
        row += 1
        worksheet.cell(row=row, column=col).value = lst[key]
    workbook.save(fname);


def create_workbook(filename, sheetname):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet(sheetname)
    workbook.close()


def test_pred_acc(dset, dsetn, dir, row, col, f_cols, title, window):
    X,Y = select_features(dset, f_cols)
    Y_pred = get_predictions(dset, 'Application', f_cols, window)
    lst = [metrics.accuracy_score(Y, Y_pred), metrics.precision_score(Y, Y_pred), metrics.recall_score(Y, Y_pred)]
    lst.extend(f_cols)
    write_to_excel1(dir + "/" + dsetn + ".xlsx", "test_outcome", row, col, Y_pred, title)
    with open(dir + "/acc_" + dsetn + ".csv", "a+") as s:
         s = csv.writer(s, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
         s.writerow(lst)

def get_kalman():
    ls_filter = KalmanFilter(dim_x=2, dim_z=1)
    ls_filter.x = np.array([0., 0.]) # intial states: memory pressure, memory pressure gradient
    ls_filter.F = np.array([[1., 1.], [0.,1.]]) # state transition matrix
    ls_filter.H = np.array([[1., 0.]]) # measurement function
    ls_filter.P *= 0. # initial uncertainty of states
    ls_filter.R = 1 # measurement noise (larger, smoothier)
    # refelcts uncertainly from higher-order unknow input (model noise)
    ls_filter.Q = Q_discrete_white_noise(dim=2, dt=0.5, var=0.1)
    return ls_filter

def apply_kalman(dataset, label_idx, val_idx):
    kal_filter = get_kalman() 
    vals = dataset[val_idx].values.tolist()
    labl = dataset[label_idx].values.tolist()   
    kal_out = []
    for x in vals:
        kal_filter.predict()
        kal_filter.update(x)
        fval = kal_filter.x[0]
        kal_out.append(fval)    
    n_lbl = val_idx + "_kalman"
    dataset[n_lbl] = kal_out
    return dataset

def preprocess_data(dataset, window, window2):
    dataset = convert_to_numeric(dataset, float_indexes)
    f_cols=[]
    for i in all_cols:
        if i != 'Label' and i != 'Application':
            f_cols.append(i)
            dataset = compute_fn_sliding_window(dataset, 'Application', i , window)
            dataset = apply_kalman(dataset, 'Application', i)   
    for i in f_cols:
        for k in functions.keys():
            dataset = compute_fn_sliding_window(dataset, 'Application', i + "_" + k + "_" + str(window) , window)
    return dataset

 
def get_apply_model(dataset, f_cols, nf, dir, window):
    if nf == 3:
         fig = plt.gcf()
         axs = fig.add_subplot(111)
         plot_3d(dataset, f_cols[0], f_cols[1], f_cols[2])
         fig.savefig(dir + '/3d_plot.png', format='png')
    else:
         fig = plt.gcf()
         axs = fig.add_subplot(111)
         p1 = sns.relplot(y=f_cols[0], x=f_cols[1], hue="Label", data=dataset,ax=axs)
         plt.close(p1.fig)
         fig.savefig(dir + '/2d_plot.png', format='png')
     
    plt.close()
    fig = plt.gcf()
    axs1 = fig.add_subplot(111)
    features, outcome = select_features(dataset, f_cols)         
    predictions = get_predictions(dataset, 'Application', f_cols, window)
    print(len(predictions) , len(outcome))
    cnf_matrix, res = verify_correctness(outcome, predictions, dir)
    plot_conf_matrix(cnf_matrix, axs1, fig)
    fig.savefig(dir + '/conf_matrix.png', format='png')
    plt.close()
    output = res + f_cols
    return output

if __name__ == "__main__":
    main()
