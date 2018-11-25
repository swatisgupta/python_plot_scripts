#!/bin/python

import pandas as pd
import xlsxwriter as xs
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
#from PIL import Image
import os.path
import sys

Images_VM = ["Virtual Memory High Water Mark.png", "Boxplots: Virtual Memory High Water Mark.png", 
	     "Consecutive Counts: Virtual Memory High Water Mark part 1.png",
             "Consecutive Counts: Virtual Memory High Water Mark part 2.png"]

Images_RSS = ["Resident Set Size.png", "Boxplots: Resident Set Size.png", "Consecutive Counts: Resident Set Size part 1.png", "Consecutive Counts: Resident Set Size part 2.png"]

Images_L1 = ["Percentage of L1 Cache misses wrt total LD & SR operations.png", "Boxplots: Percentage of L1 Cache misses wrt total LD & SR operations.png"]

Images_L2 = ["Percentage of L1 Cache misses wrt total LD & SR operations.png", "Boxplots: Percentage of L2 Cache misses wrt total LD & SR operations.png" ]

Images_L3 = [ "Percentage of L3 Cache misses wrt total LD & SR operations.png", "Boxplots: Percentage of L3 Cache misses wrt total LD & SR operations.png" ]

Images_L1_B = [ "Bandwidth L1.png" ]

Images_L2_B = [ "Bandwidth L2.png" ]

Images_L3_B = [ "Bandwidth L3.png" ]

row = [15, 46, 75, 105]

VmSheet = "Virtual Memory High Water Mark"
rssSheet = "Resident Set Size"
l1Sheet = "L1 Cache Misses"
l2Sheet = "L2 Cache Misses"
l3Sheet = "L3 Cache Misses"
l1Bandwidth = "L1 Bandwidth"
l2Bandwidth = "L2 Bandwidth"
l3Bandwidth = "L3 Bandwidth"

def create_workbook(filename):
    workbook = xs.Workbook(filename)
    worksheet = workbook.add_worksheet(VmSheet)
    worksheet = workbook.add_worksheet(rssSheet)
    worksheet = workbook.add_worksheet(l1Sheet)
    worksheet = workbook.add_worksheet(l2Sheet)
    worksheet = workbook.add_worksheet(l3Sheet)
    worksheet = workbook.add_worksheet(l1Bandwidth)
    worksheet = workbook.add_worksheet(l2Bandwidth)
    worksheet = workbook.add_worksheet(l3Bandwidth)
    workbook.close()


def write_to_excel(fname, sheetname, Cell, img):
    print(sheetname)
    print(fname)
    #workbook = xs.Workbook(filename)
    workbook = load_workbook(filename = fname)
    #worksheet = workbook.get_worksheet_by_name(sheetname)
    worksheet = workbook[sheetname]
    #worksheet.insert_image(Cell, img)
    Img = Image(img)
    print("Adding " + img + " to " + Cell)
    worksheet.add_image(Img, Cell);
    
    #workbook.close()
    workbook.save(fname);

def main():

    if len(sys.argv) != 4:
       print("Usage: python " + sys.argv[0] + " filename column image_dir")
       sys.exit()

    filename = sys.argv[1]
    Column = sys.argv[2] 
    Image_dir = sys.argv[3]
    
    if os.path.isfile(filename) == False:
        create_workbook(filename)      

    global VmSheet
    global rssSheet
    global l1Sheet
    global l2Sheet
    global l3Sheet
    for i in range(len(Images_VM)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, VmSheet, Cell, Image_dir + "/" + Images_VM[i])
    
    for i in range(len(Images_RSS)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, rssSheet, Cell, Image_dir + "/" + Images_RSS[i])
    
    for i in range(len(Images_L1)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, l1Sheet, Cell, Image_dir + "/" + Images_L1[i])
    
    for i in range(len(Images_L2)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, l2Sheet, Cell, Image_dir + "/" + Images_L2[i])

    for i in range(len(Images_L3)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, l3Sheet, Cell, Image_dir + "/" + Images_L3[i])

    for i in range(len(Images_L1_B)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, l1Bandwidth, Cell, Image_dir + "/" + Images_L1_B[i])

    for i in range(len(Images_L2_B)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, l2Bandwidth, Cell, Image_dir + "/" + Images_L2_B[i])

    for i in range(len(Images_L3_B)):
        Cell = Column + str(row[i]) 
        write_to_excel(filename, l3Bandwidth, Cell, Image_dir + "/" + Images_L3_B[i])
    


if __name__ == "__main__":
    main()	 
